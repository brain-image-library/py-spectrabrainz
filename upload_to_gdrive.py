#!/usr/bin/env python3
"""
upload_to_gdrive.py

Copyright (c) 2026 Pittsburgh Supercomputing Center (PSC),
Brain Image Library (BIL)

Author: icaoberg

Description:
    Generate an Excel report from YYYYMMDD.tsv files and upload it to Google Drive
    using rclone.

    The script performs the following steps:
        1. Find input TSV files matching the pattern YYYYMMDD.tsv
        2. Create or update 'spectrabrainz-report.xlsx'
        3. For each TSV file:
            - Load data into a worksheet named after the date
            - Sort rows by the 'completion' column (descending), if present
        4. Apply status-based row coloring using the 'state' column:
            - Completed -> green
            - Failed    -> red
            - Canceled  -> yellow
        5. Autosize all columns for readability
        6. Upload the resulting spreadsheet to Google Drive via rclone

Usage:
    python ./upload_to_gdrive.py

Requirements:
    - Python 3
    - pandas
    - openpyxl
    - rclone installed and configured
    - TSV files named like YYYYMMDD.tsv in the working directory

Notes:
    - The remote path is configured via RCLONE_REMOTE_PATH.
    - Existing sheets with the same name will be replaced.
"""

import re
import subprocess
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

RCLONE_REMOTE_PATH = "PSC:Brain_Image_Library/spectrabrainz/"


# -----------------------------------------------------------
# Find input *.tsv files
# -----------------------------------------------------------
def find_tsv_files(base_dir: Path):
    """Return a sorted list of TSV files matching YYYYMMDD.tsv in base_dir."""
    pattern = re.compile(r"^\d{8}\.tsv$")
    return sorted(
        f for f in base_dir.iterdir()
        if f.is_file() and pattern.match(f.name)
    )


# -----------------------------------------------------------
# Write Excel from TSV (pandas 2.x safe)
# -----------------------------------------------------------
def write_excel_from_tsv(tsv_files, excel_path: Path):
    """Create or update an Excel workbook from a list of TSV files."""
    if excel_path.exists():
        print(f"Appending to existing Excel file: {excel_path}")
        mode = "a"
    else:
        print(f"Creating new Excel file: {excel_path}")
        mode = "w"

    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode=mode,
        if_sheet_exists="replace",
    ) as writer:
        for tsv_file in tsv_files:
            sheet_name = tsv_file.stem
            print(f"Processing {tsv_file.name} → sheet '{sheet_name}'")

            df = pd.read_csv(tsv_file, sep="\t")

            # ---- SORT by completion (date) DESC if present
            if "completion" in df.columns:
                completion_dt = pd.to_datetime(df["completion"], errors="coerce")
                df = df.assign(_completion_dt=completion_dt).sort_values(
                    by="_completion_dt",
                    ascending=False,
                    na_position="last",
                ).drop(columns=["_completion_dt"])

                # Write back in a consistent string format (Excel-friendly),
                # keeping blanks where parsing failed.
                df["completion"] = completion_dt.dt.strftime("%Y-%m-%d %H:%M:%S").where(
                    completion_dt.notna(), ""
                )
            else:
                print("  Note: column 'completion' not found; skipping sort.")

            df.to_excel(writer, sheet_name=sheet_name, index=False)


# -----------------------------------------------------------
# Autosize columns
# -----------------------------------------------------------
def autosize_columns(ws):
    """Autosize all columns in an openpyxl worksheet based on content length."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0

        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            val = "" if val is None else str(val)
            max_len = max(max_len, len(val))

        ws.column_dimensions[letter].width = max_len + 2


# -----------------------------------------------------------
# Ensure workbook exists
# -----------------------------------------------------------
def ensure_workbook_exists(excel_path: Path):
    """Create an empty workbook if the Excel file does not yet exist."""
    if not excel_path.exists():
        print(f"Excel file missing; creating: {excel_path}")
        wb = Workbook()
        wb.save(excel_path)


# -----------------------------------------------------------
# Apply formatting (colors) + autosize
# -----------------------------------------------------------
def apply_backup_status_formatting(excel_path: Path, sheet_names):
    """Apply row coloring based on 'state' column and autosize columns."""
    print("Applying formatting and autosizing columns...")

    ensure_workbook_exists(excel_path)
    wb = load_workbook(excel_path)

    fills = {
        "Completed": PatternFill(start_color="228B22", end_color="228B22", fill_type="solid"),
        "Failed": PatternFill(start_color="B22222", end_color="B22222", fill_type="solid"),
        "Canceled": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    }

    white_font = Font(color="FFFFFF")
    black_font = Font(color="000000")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' missing; skipping.")
            continue

        ws = wb[sheet_name]
        print(f"Formatting sheet '{sheet_name}'")

        # Locate 'state' column
        state_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "state":
                state_col = col
                break

        if state_col is None:
            print("  Warning: column 'state' not found; skipping coloring.")
            autosize_columns(ws)
            continue

        for row in range(2, ws.max_row + 1):
            state = ws.cell(row=row, column=state_col).value
            fill = fills.get(state)

            if fill:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.font = black_font if state == "Canceled" else white_font

        autosize_columns(ws)

    wb.save(excel_path)
    print("Formatting + autosizing complete.")


# -----------------------------------------------------------
# rclone upload
# -----------------------------------------------------------
def upload_with_rclone(excel_path: Path, remote_path: str):
    """Upload the Excel file to the given rclone remote path."""
    print(f"Sorting done. Uploading '{excel_path}' → '{remote_path}'")
    try:
        subprocess.run(
            ["rclone", "copy", str(excel_path), remote_path, "--progress"],
            check=True,
        )
        print("Upload OK.")
    except FileNotFoundError:
        print("ERROR: rclone not found.")
    except subprocess.CalledProcessError as e:
        print(f"rclone failed: exit {e.returncode}")


# -----------------------------------------------------------
# Main
# -----------------------------------------------------------
def main():
    """Main entry point for report generation and upload."""
    cwd = Path(".").resolve()
    excel_path = cwd / "spectrabrainz-report.xlsx"

    tsv_files = find_tsv_files(cwd)
    if not tsv_files:
        print("No YYYYMMDD.tsv files found.")
        return

    sheet_names = [f.stem for f in tsv_files]

    write_excel_from_tsv(tsv_files, excel_path)
    apply_backup_status_formatting(excel_path, sheet_names)
    upload_with_rclone(excel_path, RCLONE_REMOTE_PATH)


if __name__ == "__main__":
    main()